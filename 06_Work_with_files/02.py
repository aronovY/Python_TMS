import json
import openpyxl


"""This program improves the excel table and converts its data into a .json file."""


def excel_change(file):

    wb = openpyxl.load_workbook(filename=file)

    sheet = wb.worksheets[0]  # Save the worksheet with which we will work in a variable

    col = sheet.max_column  # Get maximum columns
    row = sheet.max_row  # Get maximum rows

    sheet.insert_cols(1)

    for j in range(1, row + 1):  # Here we merge two columns
        sheet['B{}'.format(j)].value = str(sheet.cell(j, 2).value) + ' ' + str(sheet.cell(j, 3).value)
        sheet.merge_cells('B{}:C{}'.format(j, j))

    lists_marks = []

    for i in range(1, row + 1):  # In this cycle we consider the average mark
        count = 0
        list = []
        for j in range(4, col + 1):
            if sheet.cell(i, j).value == 'end':
                break
            if sheet.cell(i, j).value is None:
                continue
            else:
                list.append(int(sheet.cell(i, j).value))
                count += 1
        sheet[f'A{i}'] = '=SUM(B{}:Z{})/{}'.format(i, i, count)
        lists_marks.append(list)

    inx = 0

    res = {}

    for row_index in range(1, row + 1):
        res[sheet.cell(row_index, 2).value] = {
                        'marks': lists_marks[inx],
                        'average_marks': sum(lists_marks[inx]) / len(lists_marks[inx])
                    }

        inx += 1

    result = {'students': res}

    return result


def from_excel_to_json(rs, p):
    with open(p, 'w') as json_file:
        json_str = json.dumps(rs, indent=1)
        json_file.write(json_str)


if __name__ == '__main__':
    path_excel = 'Student.xlsx'
    path_json = 'excel_json.json'
    re = excel_change(path_excel)
    from_excel_to_json(re, path_json)
